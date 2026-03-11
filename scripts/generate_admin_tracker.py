#!/usr/bin/env python3
"""
Verseras Labs Admin Tracker — Excel Workbook Generator

Generates a professionally styled Excel workbook with:
  1. Project Plan (~50 tasks across 9 categories)
  2. Expenses (Form 1120 aligned, 28 categories)
  3. Revenue (ParlAid revenue model)
  4. Dashboard (P&L, monthly summary, Form 1120 deduction summary)
  5. Forecast vs Actual (bridge to ParlAid Financial Forecast)
  6. _Ref (hidden lookup table)

Output: OneDrive → Verseras Labs/Admin/Verseras_Labs_Admin_Tracker.xlsx
"""

import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.page import PrintPageSetup

# ─── Constants ────────────────────────────────────────────────────────────────

BRAND_BLUE = "2563EB"
BRAND_BLUE_LIGHT = "DBEAFE"
HEADER_FONT_COLOR = "FFFFFF"

TAB_COLORS = {
    "Project Plan": BRAND_BLUE,
    "Expenses": "16A34A",
    "Revenue": "F59E0B",
    "Dashboard": "7C3AED",
    "Forecast vs Actual": "0D9488",
    "_Ref": "6B7280",
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

HEADER_FILL = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color=HEADER_FONT_COLOR, size=11)
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=BRAND_BLUE)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=BRAND_BLUE)

ALT_ROW_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
PURPLE_LIGHT_FILL = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
BLUE_LIGHT_FILL = PatternFill(start_color=BRAND_BLUE_LIGHT, end_color=BRAND_BLUE_LIGHT, fill_type="solid")

CURRENCY_FMT = '#,##0.00'
DATE_FMT = "YYYY-MM-DD"
PCT_FMT = "0.0%"

OUTPUT_DIR = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Documents/Side_Business/Verseras Labs/Admin"
)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Verseras_Labs_Admin_Tracker.xlsx")

# ─── Reference Data ──────────────────────────────────────────────────────────

EXPENSE_CATEGORIES = [
    ("Compensation of Officers", "Line 12", 1.0),
    ("Salaries & Wages", "Line 13", 1.0),
    ("Repairs & Maintenance", "Line 14", 1.0),
    ("Bad Debts", "Line 15", 1.0),
    ("Rents", "Line 16", 1.0),
    ("Taxes & Licenses", "Line 17", 1.0),
    ("Interest", "Line 18", 1.0),
    ("Charitable Contributions", "Line 19", 1.0),
    ("Depreciation", "Line 20", 1.0),
    ("Pension/Profit-Sharing", "Line 23", 1.0),
    ("Employee Benefits", "Line 24", 1.0),
    ("Advertising & Marketing", "Line 26", 1.0),
    ("Legal & Professional Fees", "Line 26", 1.0),
    ("Office Expenses", "Line 26", 1.0),
    ("Supplies", "Line 26", 1.0),
    ("Travel", "Line 26", 1.0),
    ("Meals (50% Deductible)", "Line 26", 0.5),
    ("Utilities & Internet", "Line 26", 1.0),
    ("Insurance", "Line 26", 1.0),
    ("Software & Subscriptions", "Line 26", 1.0),
    ("Cloud Infrastructure", "Line 26", 1.0),
    ("Bank & Processing Fees", "Line 26", 1.0),
    ("Startup Costs (Sec 195)", "Line 26", 1.0),
    ("Organizational Costs (Sec 248)", "Line 26", 1.0),
    ("Mobile App Development", "Line 26", 1.0),
    ("Patent & IP Costs", "Line 26", 1.0),
    ("Licensing & Compliance", "Line 26", 1.0),
    ("Other Deductions", "Line 26", 1.0),
]

REVENUE_CATEGORIES = [
    "CPA Revenue",
    "Repeat-Click Revenue",
    "Subscription Revenue",
    "Consulting Revenue",
    "Licensing/Royalties",
    "Interest Income",
    "Other Income",
]

REVENUE_STATUSES = ["Invoiced", "Paid", "Partial", "Overdue", "Written Off"]

PROJECT_CATEGORIES = [
    "Formation & Structure",
    "Governance",
    "IP & Legal",
    "Tax & Compliance",
    "Banking & Finance",
    "Investor Prep",
    "Product & Operations",
    "HR & Payroll",
    "Insurance",
]

PRIORITIES = ["Critical", "High", "Medium", "Low"]
STATUSES = ["Not Started", "In Progress", "Completed", "Blocked", "Deferred"]
ENTITIES = ["Verseras Labs", "ParlAid"]
PAYMENT_METHODS = ["ACH", "Wire", "Credit Card", "Check", "Cash", "PayPal", "Stripe", "Other"]

# Pre-populated project plan tasks
# (Task, Category, Priority, Status, Due Date str, Owner, Notes)
TASKS = [
    # Formation & Structure
    ("Apply for EIN — Verseras Labs Inc.", "Formation & Structure", "Critical", "Not Started", "2025-04-01", "George", "IRS Form SS-4, online or fax"),
    ("Apply for EIN — ParlAid LLC", "Formation & Structure", "Critical", "Not Started", "2025-04-01", "George", "Separate EIN for subsidiary"),
    ("Confirm registered agent (DE)", "Formation & Structure", "High", "Not Started", "2025-04-15", "George", "Annual fee ~$50–150"),
    ("Foreign qualification in home state", "Formation & Structure", "High", "Not Started", "2025-04-30", "George", "Required if operating outside DE"),
    ("Complete Clerky post-incorporation steps", "Formation & Structure", "High", "Not Started", "2025-04-15", "George", "Action items, stock issuance, consents"),
    ("File Certificate of Formation — ParlAid LLC", "Formation & Structure", "Critical", "Not Started", "2025-04-01", "George", "DE Division of Corporations"),
    ("Draft Operating Agreement — ParlAid LLC", "Formation & Structure", "High", "Not Started", "2025-04-15", "George", "Single-member LLC (Verseras Labs as member)"),
    # Governance
    ("Adopt bylaws", "Governance", "Critical", "Not Started", "2025-04-01", "George", "Board resolution required"),
    ("Hold organizational board meeting", "Governance", "Critical", "Not Started", "2025-04-01", "George", "Approve bylaws, officers, stock, fiscal year"),
    ("Elect officers (CEO, Secretary, Treasurer)", "Governance", "Critical", "Not Started", "2025-04-01", "George", "Board resolution"),
    ("Authorize and issue founder stock", "Governance", "Critical", "Not Started", "2025-04-07", "George", "Par value, vesting schedule, stock purchase agreement"),
    ("File 83(b) election with IRS", "Governance", "Critical", "Not Started", "2025-04-10", "George", "MUST file within 30 days of stock grant"),
    ("Set up stock ledger", "Governance", "High", "Not Started", "2025-04-15", "George", "Track all equity issuances"),
    ("Create corporate minute book", "Governance", "Medium", "Not Started", "2025-05-01", "George", "Bylaws, resolutions, stock certificates"),
    ("Adopt equity incentive plan", "Governance", "Medium", "Not Started", "2025-06-01", "George", "Option pool for future employees/advisors"),
    # IP & Legal
    ("Execute IP Assignment (George → Verseras Labs)", "IP & Legal", "Critical", "Not Started", "2025-04-07", "George", "All prior IP, code, algorithms"),
    ("Execute IP License (Verseras Labs → ParlAid)", "IP & Legal", "Critical", "Not Started", "2025-04-15", "George", "Royalty/licensing terms, arm's-length pricing"),
    ("File provisional patent application", "IP & Legal", "High", "Not Started", "2025-05-15", "George", "Risk algorithm, data pipeline innovations"),
    ("File non-provisional patent (12-mo deadline)", "IP & Legal", "High", "Not Started", "2026-05-15", "George", "Must file within 12 months of provisional"),
    ("File trademark application — ParlAid", "IP & Legal", "Medium", "Not Started", "2025-06-01", "George", "USPTO, word mark + logo"),
    ("Conduct FTO (freedom-to-operate) analysis", "IP & Legal", "Medium", "Not Started", "2025-07-01", "George", "Review competitor patents"),
    ("Draft PIIA template", "IP & Legal", "Medium", "Not Started", "2025-05-01", "George", "Proprietary Info & Inventions Assignment for contractors/employees"),
    ("Engage patent attorney", "IP & Legal", "High", "Not Started", "2025-04-15", "George", "Provisonal + non-provisional strategy"),
    # Tax & Compliance
    ("Pay DE franchise tax (annual, Mar 1)", "Tax & Compliance", "Critical", "Not Started", "2026-03-01", "George", "Minimum $400/yr for C Corp"),
    ("File DE Annual Report", "Tax & Compliance", "High", "Not Started", "2026-03-01", "George", "Due with franchise tax"),
    ("File Form 1120 (federal corporate tax)", "Tax & Compliance", "Critical", "Not Started", "2026-04-15", "George", "Calendar year filer"),
    ("Pay estimated quarterly taxes (Q1)", "Tax & Compliance", "High", "Not Started", "2025-04-15", "George", "Form 1120-W"),
    ("Pay estimated quarterly taxes (Q2)", "Tax & Compliance", "High", "Not Started", "2025-06-15", "George", "Form 1120-W"),
    ("Pay estimated quarterly taxes (Q3)", "Tax & Compliance", "High", "Not Started", "2025-09-15", "George", "Form 1120-W"),
    ("Pay estimated quarterly taxes (Q4)", "Tax & Compliance", "High", "Not Started", "2026-01-15", "George", "Form 1120-W"),
    ("Register for state income tax (home state)", "Tax & Compliance", "High", "Not Started", "2025-05-01", "George", "If applicable"),
    ("Make Sec 195 startup cost election", "Tax & Compliance", "Medium", "Not Started", "2026-04-15", "George", "Deduct up to $5K in Year 1, amortize remainder over 180 months"),
    ("Make Sec 248 organizational cost election", "Tax & Compliance", "Medium", "Not Started", "2026-04-15", "George", "Deduct up to $5K in Year 1, amortize remainder over 180 months"),
    ("Choose accounting method (cash vs accrual)", "Tax & Compliance", "High", "Not Started", "2025-04-15", "George", "Cash method likely OK for startup"),
    ("Evaluate sales tax registration needs", "Tax & Compliance", "Low", "Not Started", "2025-06-01", "George", "SaaS may trigger economic nexus"),
    # Banking & Finance
    ("Open business bank account — Verseras Labs", "Banking & Finance", "Critical", "Not Started", "2025-04-07", "George", "Need EIN, Articles of Incorporation"),
    ("Open business bank account — ParlAid LLC", "Banking & Finance", "High", "Not Started", "2025-04-15", "George", "Separate from parent entity"),
    ("Apply for business credit card", "Banking & Finance", "Medium", "Not Started", "2025-04-30", "George", "Build business credit"),
    ("Set up bookkeeping system", "Banking & Finance", "High", "Not Started", "2025-04-15", "George", "QuickBooks Online or Wave"),
    ("Configure chart of accounts (Form 1120 aligned)", "Banking & Finance", "High", "Not Started", "2025-04-15", "George", "Match expense categories to tax lines"),
    # Investor Prep
    ("Finalize pitch deck", "Investor Prep", "High", "Not Started", "2025-05-01", "George", "Problem, solution, market, traction, team, ask"),
    ("Update financial model / forecast", "Investor Prep", "High", "Not Started", "2025-05-01", "George", "Align with ParlAid Financial Forecast"),
    ("Prepare SAFE / convertible note documents", "Investor Prep", "Medium", "Not Started", "2025-05-15", "George", "YC SAFE template, post-money"),
    ("Set up data room (Google Drive / DocSend)", "Investor Prep", "Medium", "Not Started", "2025-05-15", "George", "Corporate docs, financials, IP, cap table"),
    ("Build target investor list", "Investor Prep", "Medium", "Not Started", "2025-05-01", "George", "Angels, pre-seed funds, sports/gaming VCs"),
    ("Prepare due diligence package", "Investor Prep", "Low", "Not Started", "2025-06-01", "George", "Corporate docs, financials, IP assignments, cap table"),
    # Product & Operations
    ("Set up custom domain + email", "Product & Operations", "High", "Not Started", "2025-04-07", "George", "Google Workspace or Microsoft 365"),
    ("Research business license requirements", "Product & Operations", "Medium", "Not Started", "2025-05-01", "George", "City/county/state business licenses"),
    ("Draft contractor agreement template", "Product & Operations", "Medium", "Not Started", "2025-05-01", "George", "IP assignment, confidentiality, payment terms"),
    ("Set up project management tools", "Product & Operations", "Medium", "Not Started", "2025-04-15", "George", "Linear, Notion, or GitHub Projects"),
    ("Mobile app development kickoff", "Product & Operations", "High", "Not Started", "2025-09-01", "George", "Month 6 per ParlAid forecast — $60K budget"),
    # HR & Payroll
    ("Set up payroll system", "HR & Payroll", "Medium", "Not Started", "2025-06-01", "George", "Gusto, Rippling, or ADP"),
    ("Prepare W-4 / I-9 process", "HR & Payroll", "Low", "Not Started", "2025-06-01", "George", "For first hire"),
    ("Register for unemployment insurance", "HR & Payroll", "Low", "Not Started", "2025-06-01", "George", "State-specific requirement"),
    ("Draft employee handbook", "HR & Payroll", "Low", "Not Started", "2025-07-01", "George", "At-will employment, policies, benefits"),
    # Insurance
    ("Obtain general liability insurance", "Insurance", "Medium", "Not Started", "2025-05-01", "George", "Basic business coverage"),
    ("Obtain professional liability / E&O insurance", "Insurance", "Medium", "Not Started", "2025-05-01", "George", "Critical for SaaS/consulting"),
    ("Evaluate D&O insurance", "Insurance", "Low", "Not Started", "2025-06-01", "George", "Required before fundraising"),
    ("Evaluate workers' comp requirements", "Insurance", "Low", "Not Started", "2025-07-01", "George", "State-specific, needed with first employee"),
]

FORECAST_ROWS = [
    ("Cloud Infrastructure", 500, "Scales ~$150/1K MAU per ParlAid forecast"),
    ("Marketing & Growth", 1000, "8% MoM growth target"),
    ("Licensing & Compliance", 3333, "$40K/yr amortized monthly"),
    ("Legal & Patent", 5000, "$15K Year 1, concentrated M1-M3"),
    ("Mobile App Development", 8571, "$60K total, M6-M12 per forecast"),
    ("Software & Subscriptions", 300, "Dev tools, project management, analytics"),
    ("Misc / Buffer", 500, "Contingency"),
]

# Map forecast categories to Expenses sheet categories for SUMPRODUCT lookups
FORECAST_EXPENSE_MAP = {
    "Cloud Infrastructure": "Cloud Infrastructure",
    "Marketing & Growth": "Advertising & Marketing",
    "Licensing & Compliance": "Licensing & Compliance",
    "Legal & Patent": "Legal & Professional Fees",
    "Mobile App Development": "Mobile App Development",
    "Software & Subscriptions": "Software & Subscriptions",
    "Misc / Buffer": "Other Deductions",
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def style_header_row(ws, num_cols, row=1):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_area(ws, num_cols, start_row, end_row):
    """Apply alternating rows and borders to data area."""
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL


def set_column_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def setup_print(ws, landscape=True):
    """Configure print settings."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"


def add_dropdown(ws, options, col_letter, min_row, max_row, prompt_title="", prompt_body=""):
    """Add a dropdown data validation to a column range."""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = prompt_body
    dv.promptTitle = prompt_title
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid Entry"
    dv.error = f"Please select from the dropdown list."
    dv.sqref = f"{col_letter}{min_row}:{col_letter}{max_row}"
    ws.add_data_validation(dv)


# ─── Sheet Builders ───────────────────────────────────────────────────────────

def build_ref_sheet(wb):
    """Sheet 6: _Ref — hidden lookup table."""
    ws = wb.create_sheet("_Ref")
    ws.sheet_state = "hidden"
    ws.sheet_properties.tabColor = TAB_COLORS["_Ref"]

    headers = ["Category", "Form 1120 Line", "Deductibility %"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))

    for i, (cat, line, ded) in enumerate(EXPENSE_CATEGORIES, 2):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=line)
        ws.cell(row=i, column=3, value=ded)
        ws.cell(row=i, column=3).number_format = PCT_FMT

    style_data_area(ws, len(headers), 2, len(EXPENSE_CATEGORIES) + 1)
    set_column_widths(ws, {"A": 35, "B": 18, "C": 18})
    return ws


def build_project_plan(wb):
    """Sheet 1: Project Plan."""
    ws = wb.create_sheet("Project Plan", 0)
    ws.sheet_properties.tabColor = TAB_COLORS["Project Plan"]

    headers = [
        "Task", "Category", "Priority", "Status", "Due Date",
        "Completed Date", "Owner", "Blocked By", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))

    # Populate tasks
    for i, (task, cat, pri, status, due, owner, notes) in enumerate(TASKS, 2):
        ws.cell(row=i, column=1, value=task)
        ws.cell(row=i, column=2, value=cat)
        ws.cell(row=i, column=3, value=pri)
        ws.cell(row=i, column=4, value=status)
        ws.cell(row=i, column=5, value=due)
        ws.cell(row=i, column=5).number_format = DATE_FMT
        # col 6 = Completed Date (blank)
        ws.cell(row=i, column=7, value=owner)
        # col 8 = Blocked By (blank)
        ws.cell(row=i, column=9, value=notes)

    last_data_row = len(TASKS) + 1
    max_row = 200  # extend dropdowns for future rows

    style_data_area(ws, len(headers), 2, last_data_row)

    # Dropdowns
    add_dropdown(ws, PROJECT_CATEGORIES, "B", 2, max_row, "Category", "Select task category")
    add_dropdown(ws, PRIORITIES, "C", 2, max_row, "Priority", "Select priority level")
    add_dropdown(ws, STATUSES, "D", 2, max_row, "Status", "Select task status")

    # Conditional formatting — Status colors
    ws.conditional_formatting.add(
        f"D2:D{max_row}",
        CellIsRule(operator="equal", formula=['"Completed"'], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        f"D2:D{max_row}",
        CellIsRule(operator="equal", formula=['"In Progress"'], fill=AMBER_FILL)
    )
    ws.conditional_formatting.add(
        f"D2:D{max_row}",
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        f"D2:D{max_row}",
        CellIsRule(operator="equal", formula=['"Deferred"'], fill=GRAY_FILL)
    )

    # Overdue: Due Date < TODAY() AND Status != Completed
    ws.conditional_formatting.add(
        f"E2:E{max_row}",
        FormulaRule(
            formula=[f'AND(E2<TODAY(), E2<>"", D2<>"Completed")'],
            fill=RED_FILL,
            font=Font(name="Calibri", bold=True, color="DC2626")
        )
    )

    # Priority colors
    ws.conditional_formatting.add(
        f"C2:C{max_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        f"C2:C{max_row}",
        CellIsRule(operator="equal", formula=['"High"'],
                   fill=PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"))
    )

    # Column widths
    set_column_widths(ws, {
        "A": 48, "B": 22, "C": 12, "D": 14, "E": 14,
        "F": 16, "G": 12, "H": 18, "I": 40
    })

    # Freeze & filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{last_data_row}"
    setup_print(ws)

    return ws


def build_expenses(wb):
    """Sheet 2: Expenses (Form 1120 Aligned)."""
    ws = wb.create_sheet("Expenses")
    ws.sheet_properties.tabColor = TAB_COLORS["Expenses"]

    headers = [
        "Date", "Description", "Vendor/Payee", "Amount", "Payment Method",
        "Category (Form 1120)", "Form 1120 Line", "Entity",
        "Subcategory", "Tax Deductible", "Receipt", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))

    max_row = 500  # room for entries

    # VLOOKUP formula for Form 1120 Line (col G) based on Category (col F)
    # References _Ref sheet columns A:B
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=7).value = f'=IF(F{row}="","",VLOOKUP(F{row},\'_Ref\'!$A:$B,2,FALSE))'

    # Dropdowns
    cat_names = [c[0] for c in EXPENSE_CATEGORIES]
    add_dropdown(ws, cat_names, "F", 2, max_row, "Category", "Select expense category")
    add_dropdown(ws, PAYMENT_METHODS, "E", 2, max_row, "Payment Method", "Select payment method")
    add_dropdown(ws, ENTITIES, "H", 2, max_row, "Entity", "Verseras Labs or ParlAid")
    add_dropdown(ws, ["Y", "N"], "J", 2, max_row, "Tax Deductible", "Is this deductible?")
    add_dropdown(ws, ["Y", "N"], "K", 2, max_row, "Receipt", "Receipt on file?")

    # Formatting
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=1).number_format = DATE_FMT
        ws.cell(row=row, column=4).number_format = CURRENCY_FMT

    # Conditional formatting — Non-deductible (red)
    ws.conditional_formatting.add(
        f"J2:J{max_row}",
        CellIsRule(operator="equal", formula=['"N"'], fill=RED_FILL)
    )
    # Missing receipt (yellow)
    ws.conditional_formatting.add(
        f"K2:K{max_row}",
        CellIsRule(operator="equal", formula=['"N"'], fill=YELLOW_FILL)
    )
    # Meals amber highlight
    ws.conditional_formatting.add(
        f"F2:F{max_row}",
        CellIsRule(operator="equal", formula=['"Meals (50% Deductible)"'], fill=AMBER_FILL)
    )

    # Column widths
    set_column_widths(ws, {
        "A": 14, "B": 30, "C": 22, "D": 14, "E": 16,
        "F": 30, "G": 16, "H": 16, "I": 18, "J": 14,
        "K": 10, "L": 30
    })

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L1"
    setup_print(ws)

    return ws


def build_revenue(wb):
    """Sheet 3: Revenue."""
    ws = wb.create_sheet("Revenue")
    ws.sheet_properties.tabColor = TAB_COLORS["Revenue"]

    headers = [
        "Date", "Description", "Client/Source", "Amount", "Payment Method",
        "Invoice #", "Category", "Entity", "Status", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, len(headers))

    max_row = 300

    # Dropdowns
    add_dropdown(ws, REVENUE_CATEGORIES, "G", 2, max_row, "Category", "Select revenue type")
    add_dropdown(ws, PAYMENT_METHODS, "E", 2, max_row, "Payment Method", "Select payment method")
    add_dropdown(ws, ENTITIES, "H", 2, max_row, "Entity", "Verseras Labs or ParlAid")
    add_dropdown(ws, REVENUE_STATUSES, "I", 2, max_row, "Status", "Select invoice status")

    # Formatting
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=1).number_format = DATE_FMT
        ws.cell(row=row, column=4).number_format = CURRENCY_FMT

    # Conditional formatting — Overdue (red), Paid (green)
    ws.conditional_formatting.add(
        f"I2:I{max_row}",
        CellIsRule(operator="equal", formula=['"Overdue"'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        f"I2:I{max_row}",
        CellIsRule(operator="equal", formula=['"Paid"'], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        f"I2:I{max_row}",
        CellIsRule(operator="equal", formula=['"Written Off"'], fill=GRAY_FILL)
    )

    set_column_widths(ws, {
        "A": 14, "B": 30, "C": 22, "D": 14, "E": 16,
        "F": 14, "G": 22, "H": 16, "I": 14, "J": 30
    })

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J1"
    setup_print(ws)

    return ws


def build_dashboard(wb):
    """Sheet 4: Dashboard with live formulas."""
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = TAB_COLORS["Dashboard"]

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    year = date.today().year

    # ── Section 1: Monthly Expense Summary ────────────────────────────────────
    row = 1
    ws.cell(row=row, column=1, value="VERSERAS LABS — ADMIN DASHBOARD").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)

    row = 3
    ws.cell(row=row, column=1, value="Monthly Expense Summary").font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)

    row = 4
    headers = ["Category"] + months + ["YTD Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, len(headers), row=row)

    cat_names = [c[0] for c in EXPENSE_CATEGORIES]
    for i, cat in enumerate(cat_names):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=cat).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        for m in range(1, 13):
            col_idx = m + 1
            # SUMPRODUCT: sum Amount (Expenses!D) where Category (Expenses!F) matches
            # and MONTH(Date) matches and YEAR(Date) matches
            formula = (
                f'=SUMPRODUCT((Expenses!$F$2:$F$500=A{r})'
                f'*(MONTH(Expenses!$A$2:$A$500)={m})'
                f'*(YEAR(Expenses!$A$2:$A$500)={year})'
                f'*Expenses!$D$2:$D$500)'
            )
            cell = ws.cell(row=r, column=col_idx, value=formula)
            cell.number_format = CURRENCY_FMT
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
        # YTD Total
        ytd_col = 14
        first_col_letter = get_column_letter(2)
        last_col_letter = get_column_letter(13)
        ws.cell(row=r, column=ytd_col,
                value=f'=SUM({first_col_letter}{r}:{last_col_letter}{r})').number_format = CURRENCY_FMT
        ws.cell(row=r, column=ytd_col).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=r, column=ytd_col).border = THIN_BORDER

        # Alternating rows
        if i % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL

    # Total row
    total_row = row + 1 + len(cat_names)
    ws.cell(row=total_row, column=1, value="TOTAL EXPENSES").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    ws.cell(row=total_row, column=1).fill = BLUE_LIGHT_FILL
    for col_idx in range(2, 15):
        col_letter = get_column_letter(col_idx)
        first_data_row = row + 1
        last_data_row = total_row - 1
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})')
        ws.cell(row=total_row, column=col_idx).number_format = CURRENCY_FMT
        ws.cell(row=total_row, column=col_idx).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=total_row, column=col_idx).border = THIN_BORDER
        ws.cell(row=total_row, column=col_idx).fill = BLUE_LIGHT_FILL

    # ── Section 2: P&L Summary ────────────────────────────────────────────────
    pnl_start = total_row + 3
    ws.cell(row=pnl_start, column=1, value="P&L Summary").font = SECTION_FONT
    ws.merge_cells(start_row=pnl_start, start_column=1, end_row=pnl_start, end_column=14)

    pnl_header = pnl_start + 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=pnl_header, column=col, value=h)
    style_header_row(ws, len(headers), row=pnl_header)

    # Revenue row
    rev_row = pnl_header + 1
    ws.cell(row=rev_row, column=1, value="Total Revenue").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=rev_row, column=1).border = THIN_BORDER
    for m in range(1, 13):
        formula = (
            f'=SUMPRODUCT((Revenue!$I$2:$I$300<>"Written Off")'
            f'*(MONTH(Revenue!$A$2:$A$300)={m})'
            f'*(YEAR(Revenue!$A$2:$A$300)={year})'
            f'*Revenue!$D$2:$D$300)'
        )
        cell = ws.cell(row=rev_row, column=m + 1, value=formula)
        cell.number_format = CURRENCY_FMT
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
    ws.cell(row=rev_row, column=14,
            value=f'=SUM(B{rev_row}:M{rev_row})').number_format = CURRENCY_FMT
    ws.cell(row=rev_row, column=14).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=rev_row, column=14).border = THIN_BORDER

    # Expenses row (reference total expenses row)
    exp_row = rev_row + 1
    ws.cell(row=exp_row, column=1, value="Total Expenses").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=exp_row, column=1).border = THIN_BORDER
    for col_idx in range(2, 15):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=exp_row, column=col_idx,
                value=f'={col_letter}{total_row}').number_format = CURRENCY_FMT
        ws.cell(row=exp_row, column=col_idx).font = BODY_FONT
        ws.cell(row=exp_row, column=col_idx).border = THIN_BORDER

    # Net Income row
    net_row = exp_row + 1
    ws.cell(row=net_row, column=1, value="Net Income").font = Font(name="Calibri", bold=True, size=11, color=BRAND_BLUE)
    ws.cell(row=net_row, column=1).border = THIN_BORDER
    ws.cell(row=net_row, column=1).fill = BLUE_LIGHT_FILL
    for col_idx in range(2, 15):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=net_row, column=col_idx,
                value=f'={col_letter}{rev_row}-{col_letter}{exp_row}')
        ws.cell(row=net_row, column=col_idx).number_format = CURRENCY_FMT
        ws.cell(row=net_row, column=col_idx).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=net_row, column=col_idx).border = THIN_BORDER
        ws.cell(row=net_row, column=col_idx).fill = BLUE_LIGHT_FILL

    # Cumulative cash row
    cash_row = net_row + 1
    ws.cell(row=cash_row, column=1, value="Cumulative Cash Position").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=cash_row, column=1).border = THIN_BORDER
    # Jan = Net Income Jan
    ws.cell(row=cash_row, column=2, value=f'=B{net_row}').number_format = CURRENCY_FMT
    ws.cell(row=cash_row, column=2).font = BODY_FONT
    ws.cell(row=cash_row, column=2).border = THIN_BORDER
    for col_idx in range(3, 14):
        prev_letter = get_column_letter(col_idx - 1)
        col_letter = get_column_letter(col_idx)
        ws.cell(row=cash_row, column=col_idx,
                value=f'={prev_letter}{cash_row}+{col_letter}{net_row}')
        ws.cell(row=cash_row, column=col_idx).number_format = CURRENCY_FMT
        ws.cell(row=cash_row, column=col_idx).font = BODY_FONT
        ws.cell(row=cash_row, column=col_idx).border = THIN_BORDER
    ws.cell(row=cash_row, column=14, value=f'=M{cash_row}').number_format = CURRENCY_FMT
    ws.cell(row=cash_row, column=14).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=cash_row, column=14).border = THIN_BORDER

    # ── Section 3: Key Metrics ────────────────────────────────────────────────
    metrics_start = cash_row + 3
    ws.cell(row=metrics_start, column=1, value="Key Metrics").font = SECTION_FONT
    ws.merge_cells(start_row=metrics_start, start_column=1, end_row=metrics_start, end_column=4)

    metrics = [
        ("YTD Revenue", f'=N{rev_row}'),
        ("YTD Expenses", f'=N{total_row}'),
        ("Net Income", f'=N{net_row}'),
        ("Largest Expense Category",
         f'=INDEX(A{row+1}:A{total_row-1},MATCH(MAX(N{row+1}:N{total_row-1}),N{row+1}:N{total_row-1},0))'),
        ("Total Deductible Expenses",
         f'=SUMPRODUCT((Expenses!$J$2:$J$500="Y")*Expenses!$D$2:$D$500)'),
        ("Total Non-Deductible Expenses",
         f'=SUMPRODUCT((Expenses!$J$2:$J$500="N")*Expenses!$D$2:$D$500)'),
        ("Meals 50% Adjustment",
         f'=SUMPRODUCT((Expenses!$F$2:$F$500="Meals (50% Deductible)")*Expenses!$D$2:$D$500)*0.5'),
    ]
    mr = metrics_start + 1
    style_header_row(ws, 2, row=mr)
    ws.cell(row=mr, column=1, value="Metric")
    ws.cell(row=mr, column=2, value="Value")
    for i, (label, formula) in enumerate(metrics):
        r = mr + 1 + i
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        cell = ws.cell(row=r, column=2, value=formula)
        cell.border = THIN_BORDER
        if label != "Largest Expense Category":
            cell.number_format = CURRENCY_FMT
        cell.font = Font(name="Calibri", bold=True, size=11)
        if i % 2 == 1:
            ws.cell(row=r, column=1).fill = ALT_ROW_FILL
            ws.cell(row=r, column=2).fill = ALT_ROW_FILL

    # ── Section 4: Form 1120 Deduction Summary ───────────────────────────────
    f1120_start = mr + len(metrics) + 3
    ws.cell(row=f1120_start, column=1, value="Form 1120 Deduction Summary").font = SECTION_FONT
    ws.merge_cells(start_row=f1120_start, start_column=1, end_row=f1120_start, end_column=4)

    f1120_header = f1120_start + 1
    f1120_headers = ["Form 1120 Line", "Description", "YTD Amount"]
    for col, h in enumerate(f1120_headers, 1):
        ws.cell(row=f1120_header, column=col, value=h)
    style_header_row(ws, len(f1120_headers), row=f1120_header)

    # Unique Form 1120 lines
    line_items = [
        ("Line 12", "Compensation of Officers"),
        ("Line 13", "Salaries & Wages"),
        ("Line 14", "Repairs & Maintenance"),
        ("Line 15", "Bad Debts"),
        ("Line 16", "Rents"),
        ("Line 17", "Taxes & Licenses"),
        ("Line 18", "Interest"),
        ("Line 19", "Charitable Contributions"),
        ("Line 20", "Depreciation"),
        ("Line 23", "Pension/Profit-Sharing Plans"),
        ("Line 24", "Employee Benefit Programs"),
        ("Line 26", "Other Deductions (subtotal below)"),
    ]

    r = f1120_header + 1
    for i, (line, desc) in enumerate(line_items):
        ws.cell(row=r, column=1, value=line).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=desc).font = BODY_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

        if line == "Line 26":
            # Sum all Line 26 categories from Expenses
            formula = (
                f'=SUMPRODUCT((Expenses!$G$2:$G$500="{line}")'
                f'*Expenses!$D$2:$D$500)'
            )
        else:
            formula = (
                f'=SUMPRODUCT((Expenses!$G$2:$G$500="{line}")'
                f'*Expenses!$D$2:$D$500)'
            )
        cell = ws.cell(row=r, column=3, value=formula)
        cell.number_format = CURRENCY_FMT
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = THIN_BORDER

        if i % 2 == 1:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL
        r += 1

    # Line 26 breakdown
    line26_start = r + 1
    ws.cell(row=line26_start, column=1, value="Line 26 — Other Deductions Breakdown").font = Font(
        name="Calibri", bold=True, size=11, italic=True, color=BRAND_BLUE
    )
    ws.merge_cells(start_row=line26_start, start_column=1, end_row=line26_start, end_column=3)

    r = line26_start + 1
    line26_cats = [c for c in EXPENSE_CATEGORIES if c[1] == "Line 26"]
    for i, (cat, _, ded) in enumerate(line26_cats):
        ws.cell(row=r, column=1, value="").border = THIN_BORDER
        ws.cell(row=r, column=2, value=f"  {cat}").font = BODY_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        formula = (
            f'=SUMPRODUCT((Expenses!$F$2:$F$500="{cat}")'
            f'*Expenses!$D$2:$D$500)'
        )
        cell = ws.cell(row=r, column=3, value=formula)
        cell.number_format = CURRENCY_FMT
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if ded < 1.0:
            note = ws.cell(row=r, column=4, value=f"Note: {int(ded*100)}% deductible")
            note.font = Font(name="Calibri", italic=True, size=10, color="B45309")
        if i % 2 == 1:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL
        r += 1

    # Grand total
    grand_total_row = r
    ws.cell(row=grand_total_row, column=1, value="").border = THIN_BORDER
    ws.cell(row=grand_total_row, column=1).fill = BLUE_LIGHT_FILL
    ws.cell(row=grand_total_row, column=2, value="TOTAL DEDUCTIONS").font = Font(
        name="Calibri", bold=True, size=11)
    ws.cell(row=grand_total_row, column=2).border = THIN_BORDER
    ws.cell(row=grand_total_row, column=2).fill = BLUE_LIGHT_FILL
    # Sum all expenses (using the total from expense summary)
    ws.cell(row=grand_total_row, column=3,
            value=f'=SUMPRODUCT((Expenses!$J$2:$J$500="Y")*Expenses!$D$2:$D$500)')
    ws.cell(row=grand_total_row, column=3).number_format = CURRENCY_FMT
    ws.cell(row=grand_total_row, column=3).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=grand_total_row, column=3).border = THIN_BORDER
    ws.cell(row=grand_total_row, column=3).fill = BLUE_LIGHT_FILL

    # Column widths
    set_column_widths(ws, {
        "A": 35, "B": 35, "C": 16, "D": 16, "E": 14,
        "F": 14, "G": 14, "H": 14, "I": 14, "J": 14,
        "K": 14, "L": 14, "M": 14, "N": 16
    })

    ws.freeze_panes = "A2"
    setup_print(ws)

    return ws


def build_forecast_vs_actual(wb):
    """Sheet 5: Forecast vs Actual."""
    ws = wb.create_sheet("Forecast vs Actual")
    ws.sheet_properties.tabColor = TAB_COLORS["Forecast vs Actual"]

    year = date.today().year
    current_month = date.today().month

    # Title
    ws.cell(row=1, column=1, value="FORECAST vs ACTUAL — Cost Tracking").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1,
            value=f"Current Month: {date.today().strftime('%B %Y')}").font = Font(
        name="Calibri", italic=True, size=11, color="6B7280")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Headers
    row = 4
    headers = ["Cost Category", "Monthly Forecast", "Actual MTD", "Variance ($)", "Variance (%)", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, len(headers), row=row)

    for i, (cat, forecast, notes) in enumerate(FORECAST_ROWS):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=cat).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        ws.cell(row=r, column=2, value=forecast).number_format = CURRENCY_FMT
        ws.cell(row=r, column=2).font = BODY_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

        # Actual MTD — SUMPRODUCT from Expenses sheet matching the mapped category
        expense_cat = FORECAST_EXPENSE_MAP.get(cat, cat)
        actual_formula = (
            f'=SUMPRODUCT((Expenses!$F$2:$F$500="{expense_cat}")'
            f'*(MONTH(Expenses!$A$2:$A$500)=MONTH(TODAY()))'
            f'*(YEAR(Expenses!$A$2:$A$500)=YEAR(TODAY()))'
            f'*Expenses!$D$2:$D$500)'
        )
        ws.cell(row=r, column=3, value=actual_formula).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3).font = BODY_FONT
        ws.cell(row=r, column=3).border = THIN_BORDER

        # Variance ($)
        ws.cell(row=r, column=4, value=f'=B{r}-C{r}').number_format = CURRENCY_FMT
        ws.cell(row=r, column=4).font = BODY_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER

        # Variance (%)
        ws.cell(row=r, column=5,
                value=f'=IF(B{r}=0,"N/A",(B{r}-C{r})/B{r})').number_format = PCT_FMT
        ws.cell(row=r, column=5).font = BODY_FONT
        ws.cell(row=r, column=5).border = THIN_BORDER

        ws.cell(row=r, column=6, value=notes).font = Font(name="Calibri", size=10, italic=True, color="6B7280")
        ws.cell(row=r, column=6).border = THIN_BORDER

        if i % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL

    # Total row
    last_data = row + len(FORECAST_ROWS)
    total_r = last_data + 1
    ws.cell(row=total_r, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_r, column=1).border = THIN_BORDER
    ws.cell(row=total_r, column=1).fill = BLUE_LIGHT_FILL
    for col_idx in range(2, 5):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_r, column=col_idx,
                value=f'=SUM({col_letter}{row+1}:{col_letter}{last_data})')
        ws.cell(row=total_r, column=col_idx).number_format = CURRENCY_FMT
        ws.cell(row=total_r, column=col_idx).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=total_r, column=col_idx).border = THIN_BORDER
        ws.cell(row=total_r, column=col_idx).fill = BLUE_LIGHT_FILL
    ws.cell(row=total_r, column=5,
            value=f'=IF(B{total_r}=0,"N/A",(B{total_r}-C{total_r})/B{total_r})').number_format = PCT_FMT
    ws.cell(row=total_r, column=5).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_r, column=5).border = THIN_BORDER
    ws.cell(row=total_r, column=5).fill = BLUE_LIGHT_FILL
    ws.cell(row=total_r, column=6).border = THIN_BORDER
    ws.cell(row=total_r, column=6).fill = BLUE_LIGHT_FILL

    # Conditional formatting — negative variance (over budget) = red
    var_range = f"D{row+1}:D{last_data}"
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL,
                   font=Font(name="Calibri", color="DC2626"))
    )
    # Under budget = green
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
    )

    # ── YTD Comparison Section ────────────────────────────────────────────────
    ytd_start = total_r + 3
    ws.cell(row=ytd_start, column=1, value="Year-to-Date Comparison").font = SECTION_FONT
    ws.merge_cells(start_row=ytd_start, start_column=1, end_row=ytd_start, end_column=6)

    ytd_header = ytd_start + 1
    ytd_headers = ["Cost Category", "Annual Forecast", "YTD Forecast", "YTD Actual", "YTD Variance", "Notes"]
    for col, h in enumerate(ytd_headers, 1):
        ws.cell(row=ytd_header, column=col, value=h)
    style_header_row(ws, len(ytd_headers), row=ytd_header)

    for i, (cat, forecast, notes) in enumerate(FORECAST_ROWS):
        r = ytd_header + 1 + i
        ws.cell(row=r, column=1, value=cat).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER

        # Annual forecast
        ws.cell(row=r, column=2, value=forecast * 12).number_format = CURRENCY_FMT
        ws.cell(row=r, column=2).font = BODY_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

        # YTD forecast (monthly × months elapsed)
        ws.cell(row=r, column=3,
                value=f'=B{r}/12*MONTH(TODAY())').number_format = CURRENCY_FMT
        ws.cell(row=r, column=3).font = BODY_FONT
        ws.cell(row=r, column=3).border = THIN_BORDER

        # YTD actual
        expense_cat = FORECAST_EXPENSE_MAP.get(cat, cat)
        ytd_formula = (
            f'=SUMPRODUCT((Expenses!$F$2:$F$500="{expense_cat}")'
            f'*(YEAR(Expenses!$A$2:$A$500)=YEAR(TODAY()))'
            f'*Expenses!$D$2:$D$500)'
        )
        ws.cell(row=r, column=4, value=ytd_formula).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4).font = BODY_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER

        # YTD variance
        ws.cell(row=r, column=5, value=f'=C{r}-D{r}').number_format = CURRENCY_FMT
        ws.cell(row=r, column=5).font = BODY_FONT
        ws.cell(row=r, column=5).border = THIN_BORDER

        ws.cell(row=r, column=6, value="").font = BODY_FONT
        ws.cell(row=r, column=6).border = THIN_BORDER

        if i % 2 == 1:
            for c in range(1, len(ytd_headers) + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL

    set_column_widths(ws, {
        "A": 28, "B": 18, "C": 16, "D": 16, "E": 16, "F": 42
    })

    ws.freeze_panes = "A5"
    setup_print(ws)

    return ws


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets in order (ref first so it exists for VLOOKUP)
    build_ref_sheet(wb)
    build_project_plan(wb)
    build_expenses(wb)
    build_revenue(wb)
    build_dashboard(wb)
    build_forecast_vs_actual(wb)

    # Reorder sheets: Project Plan, Expenses, Revenue, Dashboard, Forecast vs Actual, _Ref
    desired_order = ["Project Plan", "Expenses", "Revenue", "Dashboard", "Forecast vs Actual", "_Ref"]
    sheet_map = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [sheet_map[name] for name in desired_order]

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Save
    wb.save(OUTPUT_FILE)
    print(f"Workbook saved to: {OUTPUT_FILE}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets if ws.sheet_state != 'hidden']}")
    print(f"Hidden: {[ws.title for ws in wb.worksheets if ws.sheet_state == 'hidden']}")
    print(f"Tasks: {len(TASKS)}")
    print(f"Expense categories: {len(EXPENSE_CATEGORIES)}")


if __name__ == "__main__":
    main()
